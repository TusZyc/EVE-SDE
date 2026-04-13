from __future__ import annotations

import json
import os
import re
import shutil
import tempfile
import urllib.request
import zipfile
from collections import Counter
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
DIST = ROOT / "dist"
TMP = ROOT / ".tmp"

LATEST_URL = "https://developers.eveonline.com/static-data/tranquility/latest.jsonl"
ZIP_TEMPLATE = "https://developers.eveonline.com/static-data/tranquility/eve-online-static-data-{build}-{variant}.zip"
CEVE_XLSX_URL = os.environ.get("CEVE_MARKET_XLSX_URL", "https://www.ceve-market.org/dumps/evedata.xlsx")

VARIANT = os.environ.get("SDE_VARIANT", "jsonl")
MAX_RECORDS = int(os.environ.get("MAX_RECORDS_PER_SHARD", "600"))
MAX_BYTES = int(os.environ.get("MAX_BYTES_PER_SHARD", "2500000"))
SEARCH_MIN = 2
USER_AGENT = "Mozilla/5.0 (compatible; EVE-SDE-Browser/2.0)"
SEARCH_INDEX_FILES = {
    "types",
    "groups",
    "categories",
    "marketGroups",
    "metaGroups",
    "dogmaAttributes",
    "dogmaEffects",
    "mapRegions",
    "mapConstellations",
    "mapSolarSystems",
    "npcStations",
    "factions",
    "races",
    "blueprints",
    "skills",
}

PREFERRED_LANGS = ("zh", "zh-cn", "en", "en-us", "ja", "de", "fr", "ko", "ru", "es")

FILE_INFO = {
    "types": {"label": "物品类型", "desc": "游戏中的具体物品、舰船、模块、技能书、蓝图等类型定义。"},
    "groups": {"label": "物品分组", "desc": "类型所属的功能分组，如护盾模块、弹药、巡洋舰等。"},
    "categories": {"label": "物品大类", "desc": "更高层级的分类，如舰船、模块、技能、行星工业等。"},
    "marketgroups": {"label": "市场分类", "desc": "市场树中的分类结构，决定物品在市场浏览器中的层级。"},
    "metagroups": {"label": "元组分类", "desc": "Meta Group，用于区分 T1、T2、势力、死亡空间、官员等。"},
    "dogmaattributes": {"label": "Dogma 属性", "desc": "数值字段定义，如射程、CPU、PG、伤害修正、速度等。"},
    "dogmaeffects": {"label": "Dogma 效果", "desc": "物品或技能附带的规则效果定义。"},
    "mapregions": {"label": "星域", "desc": "EVE 宇宙中的 Region。"},
    "mapconstellations": {"label": "星座", "desc": "EVE 宇宙中的 Constellation。"},
    "mapsolarsystems": {"label": "星系", "desc": "EVE 宇宙中的 Solar System。"},
    "npcstations": {"label": "NPC 空间站", "desc": "官方 SDE 中的 NPC 空间站静态数据。"},
    "stastations": {"label": "空间站", "desc": "NPC 空间站资料。"},
    "factions": {"label": "势力", "desc": "NPC 势力与阵营定义。"},
    "races": {"label": "种族", "desc": "艾玛、加达里、盖伦特、米玛塔尔等种族定义。"},
    "blueprints": {"label": "蓝图", "desc": "蓝图生产、发明、研究活动定义。"},
    "schematics": {"label": "行星工业配方", "desc": "PI 设施的配方与产出定义。"},
    "skills": {"label": "技能", "desc": "技能训练属性、前置技能与训练倍率。"},
    "units": {"label": "单位", "desc": "属性数值使用的单位定义。"},
    "icons": {"label": "图标", "desc": "图标资源索引。"},
}

FILE_INFO.update(
    {
        "agentsinspace": {"label": "太空代理人", "desc": "代理人在太空中的位置记录。"},
        "agenttypes": {"label": "代理人类型", "desc": "代理人类型和职业分类。"},
        "ancestries": {"label": "血统出身", "desc": "角色创建时的出身背景。"},
        "bloodlines": {"label": "血统", "desc": "四大种族下的血统资料。"},
        "certificates": {"label": "证书", "desc": "技能证书和推荐技能组合。"},
        "characterattributes": {"label": "角色属性", "desc": "感知、记忆力、毅力、智力、魅力等角色属性定义。"},
        "clonegrades": {"label": "克隆等级", "desc": "历史克隆等级资料。"},
        "compressibletypes": {"label": "可压缩物品", "desc": "可以压缩和对应压缩产物的物品关系。"},
        "contrabandtypes": {"label": "违禁品", "desc": "不同势力区域的违禁品规则。"},
        "controltowerresources": {"label": "控制塔资源", "desc": "POS 控制塔燃料和资源规则。"},
        "corporationactivities": {"label": "军团活动类型", "desc": "NPC 军团的活动类型定义。"},
        "dbuffcollections": {"label": "增减益集合", "desc": "Buff/Debuff 相关集合定义。"},
        "dogmaattributecategories": {"label": "Dogma 属性分类", "desc": "Dogma 属性的分组分类。"},
        "dogmaunits": {"label": "Dogma 单位", "desc": "Dogma 属性使用的单位定义。"},
        "dynamicitemattributes": {"label": "动态物品属性", "desc": "深渊装备等动态物品可变化属性定义。"},
        "graphics": {"label": "图形资源", "desc": "物品、星体、空间对象的图形资源索引。"},
        "landmarks": {"label": "地标", "desc": "宇宙中的地标和说明。"},
        "mapasteroidbelts": {"label": "小行星带", "desc": "星系内小行星带静态位置。"},
        "mapmoons": {"label": "卫星", "desc": "星系内卫星静态位置。"},
        "mapplanets": {"label": "行星", "desc": "星系内行星静态位置。"},
        "mapstargates": {"label": "星门", "desc": "星门和跳跃连接静态数据。"},
        "mapstars": {"label": "恒星", "desc": "恒星类型和位置数据。"},
        "metagroups": {"label": "Meta 分组", "desc": "T1、T2、势力、官员、死亡空间等 Meta 分类。"},
        "npccharacters": {"label": "NPC 角色", "desc": "NPC 角色静态定义。"},
        "npccorporations": {"label": "NPC 军团", "desc": "NPC 军团资料。"},
        "npcstations": {"label": "NPC 空间站", "desc": "官方 SDE 中的 NPC 空间站静态数据。"},
        "planetresources": {"label": "行星资源", "desc": "行星资源分布相关数据。"},
        "planetSchematics": {"label": "行星工业配方", "desc": "PI 配方和产物定义。"},
        "researchagents": {"label": "科研代理人", "desc": "科研代理人和研究领域。"},
        "skinlicenses": {"label": "涂装许可证", "desc": "SKIN 许可证物品关系。"},
        "skins": {"label": "舰船涂装", "desc": "SKIN 涂装定义。"},
        "sovereigntyupgrades": {"label": "主权升级", "desc": "主权设施升级资料。"},
        "stationoperations": {"label": "空间站运营类型", "desc": "NPC 空间站运营类型。"},
        "stationservices": {"label": "空间站服务", "desc": "空间站可提供服务定义。"},
        "tournamentrulesets": {"label": "锦标赛规则集", "desc": "锦标赛规则配置。"},
        "typematerials": {"label": "物品材料", "desc": "物品精炼或制造基础材料关系。"},
    }
)

FIELD_INFO = {
    "_key": {"label": "记录键", "meaning": "该记录在当前文件中的主键或推断键。"},
    "typeID": {"label": "物品 ID", "meaning": "游戏内物品类型的唯一编号。"},
    "groupID": {"label": "分组 ID", "meaning": "该物品所属的 Group 编号。"},
    "categoryID": {"label": "大类 ID", "meaning": "该物品所属的 Category 编号。"},
    "marketGroupID": {"label": "市场分类 ID", "meaning": "该物品在市场树中的分类编号。"},
    "metaGroupID": {"label": "Meta 分类 ID", "meaning": "物品所属的 Meta 级别分组。"},
    "metaLevel": {"label": "Meta 等级", "meaning": "物品的 Meta Level 数值。"},
    "factionID": {"label": "势力 ID", "meaning": "与该物品或地点关联的 NPC 势力编号。"},
    "raceID": {"label": "种族 ID", "meaning": "关联种族编号。"},
    "regionID": {"label": "星域 ID", "meaning": "Region 唯一编号。"},
    "constellationID": {"label": "星座 ID", "meaning": "Constellation 唯一编号。"},
    "solarSystemID": {"label": "星系 ID", "meaning": "Solar System 唯一编号。"},
    "stationID": {"label": "空间站 ID", "meaning": "NPC 空间站唯一编号。"},
    "name": {"label": "名称", "meaning": "该记录在游戏内显示的名称，通常带多语言版本。"},
    "description": {"label": "描述", "meaning": "游戏内说明文本。"},
    "published": {"label": "是否发布", "meaning": "是否为对玩家公开可见的正式内容。"},
    "mass": {"label": "质量", "meaning": "物品或舰船的质量。"},
    "volume": {"label": "体积", "meaning": "物品体积。"},
    "capacity": {"label": "容量", "meaning": "容器、货舱、模块等可容纳体积。"},
    "portionSize": {"label": "份额", "meaning": "市场与工业中使用的默认份额数量。"},
    "radius": {"label": "半径", "meaning": "空间对象半径。"},
    "basePrice": {"label": "基础价格", "meaning": "系统定义的基础价格，用于税费、爆损等计算。"},
    "graphicID": {"label": "模型 ID", "meaning": "绑定到资源模型的图形编号。"},
    "iconID": {"label": "图标 ID", "meaning": "绑定到图标资源的编号。"},
    "soundID": {"label": "音效 ID", "meaning": "关联音效编号。"},
    "parentGroupID": {"label": "父级市场分类", "meaning": "市场树中的父节点编号。"},
    "hasTypes": {"label": "含物品", "meaning": "该市场分类下是否直接挂物品。"},
    "typeIDs": {"label": "物品列表", "meaning": "该分组、分类或市场分类中关联的物品 ID 列表。"},
    "groupIDs": {"label": "分组列表", "meaning": "该分类下包含的 Group 列表。"},
    "childMarketGroupIDs": {"label": "子市场分类", "meaning": "该市场分类的子节点列表。"},
    "dogmaAttributes": {"label": "属性值", "meaning": "该物品挂载的 Dogma 属性实值。"},
    "dogmaEffects": {"label": "效果列表", "meaning": "该物品挂载的 Dogma 效果。"},
    "requiredSkills": {"label": "前置技能", "meaning": "使用或学习该内容所需的技能与等级。"},
    "traits": {"label": "加成描述", "meaning": "舰船、结构或物品的加成说明。"},
    "activities": {"label": "工业活动", "meaning": "蓝图涉及的生产、复制、发明等活动。"},
    "materials": {"label": "材料", "meaning": "制造或反应所需输入材料。"},
    "products": {"label": "产物", "meaning": "制造或反应的输出结果。"},
    "time": {"label": "耗时", "meaning": "对应活动所需时间。"},
    "wormholeClassID": {"label": "虫洞等级", "meaning": "虫洞星系所属的 W-space 等级。"},
}

FIELD_INFO.update(
    {
        "constellationIDs": {"label": "星座列表", "meaning": "该星域包含的星座 ID 列表。"},
        "solarSystemIDs": {"label": "星系列表", "meaning": "该星座包含的星系 ID 列表。"},
        "planetIDs": {"label": "行星列表", "meaning": "该星系包含的行星 ID 列表。"},
        "moonIDs": {"label": "卫星列表", "meaning": "该行星包含的卫星 ID 列表。"},
        "stargateIDs": {"label": "星门列表", "meaning": "该星系包含的星门 ID 列表。"},
        "starID": {"label": "恒星 ID", "meaning": "该星系中心恒星的 ID。"},
        "securityStatus": {"label": "安全等级", "meaning": "星系安全等级，通常显示为 -1.0 到 1.0。"},
        "securityClass": {"label": "安全等级分类", "meaning": "官方 SDE 对星系安全等级的粗分类。"},
        "luminosity": {"label": "亮度", "meaning": "恒星或星系相关的亮度参数。"},
        "position": {"label": "三维坐标", "meaning": "对象在宇宙空间中的三维坐标。"},
        "position2D": {"label": "二维坐标", "meaning": "星图使用的二维投影坐标。"},
        "x": {"label": "X 坐标", "meaning": "坐标轴 X 数值。"},
        "y": {"label": "Y 坐标", "meaning": "坐标轴 Y 数值。"},
        "z": {"label": "Z 坐标", "meaning": "坐标轴 Z 数值。"},
        "border": {"label": "边境星系", "meaning": "该星系是否位于区域边界。"},
        "corridor": {"label": "通道星系", "meaning": "该星系是否属于交通通道。"},
        "fringe": {"label": "边缘星系", "meaning": "该星系是否属于边缘区域。"},
        "hub": {"label": "枢纽星系", "meaning": "该星系是否属于交通或区域枢纽。"},
        "international": {"label": "跨势力连接", "meaning": "该星系是否涉及跨势力区域连接。"},
        "regional": {"label": "跨星域连接", "meaning": "该星系是否涉及跨星域连接。"},
        "celestialIndex": {"label": "天体序号", "meaning": "天体在星系内的序号。"},
        "orbitID": {"label": "环绕目标 ID", "meaning": "该天体或空间站所环绕的对象 ID。"},
        "orbitIndex": {"label": "轨道序号", "meaning": "围绕目标时的轨道序号。"},
        "operationID": {"label": "运营类型 ID", "meaning": "NPC 空间站运营类型 ID。"},
        "ownerID": {"label": "所有者 ID", "meaning": "拥有该对象的势力、军团或实体 ID。"},
        "useOperationName": {"label": "使用运营类型命名", "meaning": "空间站名称是否使用运营类型作为名称组成部分。"},
        "reprocessingEfficiency": {"label": "精炼效率", "meaning": "空间站基础精炼效率。"},
        "reprocessingHangarFlag": {"label": "精炼机库标记", "meaning": "精炼服务使用的机库 Flag。"},
        "reprocessingStationsTake": {"label": "精炼站税", "meaning": "空间站收取的精炼比例。"},
        "factionID": {"label": "势力 ID", "meaning": "关联 NPC 势力 ID。"},
        "corporationID": {"label": "军团 ID", "meaning": "关联 NPC 或玩家军团 ID。"},
        "allianceID": {"label": "联盟 ID", "meaning": "关联联盟 ID。"},
        "agentID": {"label": "代理人 ID", "meaning": "NPC 代理人的唯一 ID。"},
        "agentTypeID": {"label": "代理人类型 ID", "meaning": "代理人的类型 ID。"},
        "divisionID": {"label": "部门 ID", "meaning": "代理人或军团所属部门 ID。"},
        "level": {"label": "等级", "meaning": "代理人、技能或规则中使用的等级。"},
        "quality": {"label": "品质", "meaning": "旧代理人系统中的品质参数。"},
        "isLocator": {"label": "定位代理人", "meaning": "该代理人是否提供定位服务。"},
        "skillID": {"label": "技能 ID", "meaning": "技能类型的 typeID。"},
        "skillLevel": {"label": "技能等级", "meaning": "要求或提供的技能等级。"},
        "primaryAttribute": {"label": "主属性", "meaning": "技能训练使用的主角色属性。"},
        "secondaryAttribute": {"label": "副属性", "meaning": "技能训练使用的副角色属性。"},
        "trainingTimeMultiplier": {"label": "训练时间倍率", "meaning": "技能训练所需时间倍率。"},
        "attributeID": {"label": "属性 ID", "meaning": "Dogma 属性 ID。"},
        "attributeName": {"label": "属性内部名", "meaning": "Dogma 属性的内部英文名。"},
        "attributeCategoryID": {"label": "属性分类 ID", "meaning": "Dogma 属性所属分类 ID。"},
        "dataType": {"label": "数据类型", "meaning": "Dogma 属性值的数据类型。"},
        "defaultValue": {"label": "默认值", "meaning": "Dogma 属性未设置时使用的默认值。"},
        "displayName": {"label": "显示名称", "meaning": "游戏 UI 中展示的名称。"},
        "displayWhenZero": {"label": "零值时显示", "meaning": "属性为 0 时是否仍在 UI 中显示。"},
        "highIsGood": {"label": "越高越好", "meaning": "该属性数值越高是否代表效果越好。"},
        "stackable": {"label": "可堆叠", "meaning": "属性效果是否可堆叠。"},
        "unitID": {"label": "单位 ID", "meaning": "属性显示使用的单位 ID。"},
        "tooltipTitle": {"label": "提示标题", "meaning": "游戏 UI Tooltip 标题。"},
        "tooltipDescription": {"label": "提示说明", "meaning": "游戏 UI Tooltip 说明。"},
        "effectID": {"label": "效果 ID", "meaning": "Dogma 效果 ID。"},
        "effectName": {"label": "效果内部名", "meaning": "Dogma 效果的内部英文名。"},
        "effectCategoryID": {"label": "效果分类 ID", "meaning": "Dogma 效果分类。"},
        "guid": {"label": "效果路径", "meaning": "Dogma 效果的内部路径标识。"},
        "isOffensive": {"label": "攻击效果", "meaning": "该效果是否属于攻击性效果。"},
        "isAssistance": {"label": "支援效果", "meaning": "该效果是否属于支援性效果。"},
        "isWarpSafe": {"label": "跃迁安全", "meaning": "该效果是否可在跃迁相关状态下安全处理。"},
        "disallowAutoRepeat": {"label": "禁止自动循环", "meaning": "模块效果是否禁止自动重复循环。"},
        "dischargeAttributeID": {"label": "电容消耗属性 ID", "meaning": "表示电容消耗的 Dogma 属性 ID。"},
        "durationAttributeID": {"label": "持续时间属性 ID", "meaning": "表示循环时间或持续时间的 Dogma 属性 ID。"},
        "falloffAttributeID": {"label": "失准范围属性 ID", "meaning": "表示 Falloff 的 Dogma 属性 ID。"},
        "rangeAttributeID": {"label": "射程属性 ID", "meaning": "表示最佳射程或作用范围的 Dogma 属性 ID。"},
        "trackingSpeedAttributeID": {"label": "跟踪速度属性 ID", "meaning": "表示炮台跟踪速度的 Dogma 属性 ID。"},
        "electronicChance": {"label": "电子战概率", "meaning": "效果是否涉及电子战概率判定。"},
        "propulsionChance": {"label": "推进干扰概率", "meaning": "效果是否涉及推进相关概率判定。"},
        "rangeChance": {"label": "范围概率", "meaning": "效果是否涉及范围概率判定。"},
        "anchorable": {"label": "可锚定", "meaning": "对象是否可以锚定。"},
        "anchored": {"label": "已锚定", "meaning": "对象是否处于锚定状态。"},
        "fittableNonSingleton": {"label": "非单例可装配", "meaning": "非单例物品是否可以被装配。"},
        "useBasePrice": {"label": "使用基础价格", "meaning": "市场或估值逻辑是否使用基础价格。"},
        "raceID": {"label": "种族 ID", "meaning": "关联种族 ID。"},
        "bloodlineID": {"label": "血统 ID", "meaning": "角色血统 ID。"},
        "ancestryID": {"label": "出身 ID", "meaning": "角色出身背景 ID。"},
        "shortDescription": {"label": "简短描述", "meaning": "较短的说明文本。"},
        "sofFactionName": {"label": "SOF 势力名", "meaning": "图形资源系统使用的势力名称。"},
        "sofHullName": {"label": "SOF 船体名", "meaning": "图形资源系统使用的船体名称。"},
        "sofRaceName": {"label": "SOF 种族名", "meaning": "图形资源系统使用的种族名称。"},
    }
)

ID_FIELD_ENTITY_HINTS = {
    "typeID": "type",
    "groupID": "group",
    "categoryID": "category",
    "marketGroupID": "marketgroup",
    "metaGroupID": "metagroup",
    "regionID": "region",
    "constellationID": "constellation",
    "solarSystemID": "system",
    "stationID": "station",
    "factionID": "faction",
    "raceID": "race",
}

ENTITY_FILE_HINTS = {
    "types": "type",
    "groups": "group",
    "categories": "category",
    "marketgroups": "marketgroup",
    "metagroups": "metagroup",
    "mapregions": "region",
    "mapconstellations": "constellation",
    "mapsolarsystems": "system",
    "npcstations": "station",
    "stastations": "station",
    "factions": "faction",
    "races": "race",
}


def request(url: str):
    return urllib.request.urlopen(urllib.request.Request(url, headers={"User-Agent": USER_AGENT}))


def text(url: str) -> str:
    with request(url) as r:
        return r.read().decode("utf-8")


def download(url: str, dest: Path) -> None:
    with request(url) as r, dest.open("wb") as f:
        shutil.copyfileobj(r, f)


def clean(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def latest_build() -> int:
    nums: list[int] = []
    for line in text(LATEST_URL).splitlines():
        if not line.strip():
            continue
        obj = json.loads(line)
        for value in obj.values():
            if isinstance(value, int):
                nums.append(value)
            elif isinstance(value, str) and value.isdigit() and len(value) >= 6:
                nums.append(int(value))
    if not nums:
        raise RuntimeError("Could not determine latest SDE build")
    return max(nums)


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def split_words(name: str) -> list[str]:
    parts = re.findall(r"[A-Z]?[a-z]+|[A-Z]+(?![a-z])|\d+", name)
    return [p.lower() for p in parts if p]


WORD_LABELS = {
    "id": "ID",
    "ids": "ID 列表",
    "name": "名称",
    "names": "名称",
    "type": "类型",
    "types": "类型",
    "group": "分组",
    "groups": "分组",
    "category": "分类",
    "market": "市场",
    "meta": "Meta",
    "dogma": "Dogma",
    "attribute": "属性",
    "attributes": "属性",
    "effect": "效果",
    "effects": "效果",
    "region": "星域",
    "constellation": "星座",
    "solar": "恒星系",
    "system": "星系",
    "station": "空间站",
    "stations": "空间站",
    "planet": "行星",
    "moon": "卫星",
    "star": "恒星",
    "stargate": "星门",
    "security": "安全",
    "status": "状态",
    "class": "等级",
    "position": "坐标",
    "radius": "半径",
    "mass": "质量",
    "volume": "体积",
    "capacity": "容量",
    "base": "基础",
    "price": "价格",
    "published": "发布",
    "description": "描述",
    "icon": "图标",
    "graphic": "图形",
    "sound": "音效",
    "race": "种族",
    "faction": "势力",
    "corporation": "军团",
    "alliance": "联盟",
    "owner": "所有者",
    "agent": "代理人",
    "skill": "技能",
    "level": "等级",
    "required": "需求",
    "primary": "主",
    "secondary": "副",
    "time": "时间",
    "duration": "持续时间",
    "range": "范围",
    "falloff": "失准",
    "damage": "伤害",
    "shield": "护盾",
    "armor": "装甲",
    "structure": "结构",
    "online": "在线",
    "active": "激活",
    "bonus": "加成",
    "value": "值",
    "default": "默认",
    "display": "显示",
    "tooltip": "提示",
    "title": "标题",
    "unit": "单位",
    "material": "材料",
    "materials": "材料",
    "product": "产物",
    "products": "产物",
    "activity": "活动",
    "activities": "活动",
}


def with_original(label: str, original: str) -> str:
    if not original or label == original or label.endswith(f"（{original}）"):
        return label
    return f"{label}（{original}）"


def humanize_field(name: str) -> str:
    if name in FIELD_INFO:
        return with_original(FIELD_INFO[name]["label"], name)
    if name.lower() in {k.lower(): k for k in FIELD_INFO}.keys():
        for key, value in FIELD_INFO.items():
            if key.lower() == name.lower():
                return with_original(value["label"], name)
    words = split_words(name)
    translated = " ".join(WORD_LABELS.get(word, word) for word in words).strip()
    return with_original(translated or name, name)


def field_meaning(name: str) -> str:
    info = FIELD_INFO.get(name)
    if info:
        return info["meaning"]
    lowered = name.lower()
    for key, value in FIELD_INFO.items():
        if key.lower() == lowered:
            return value["meaning"]
    if lowered.endswith("id"):
        return "这是一个 ID 字段，用来关联到另一份 SDE 数据记录。"
    if lowered.endswith("ids"):
        return "这是一个 ID 列表字段，用来关联到多条 SDE 记录。"
    if lowered.endswith("name"):
        return "这是一个名称字段。"
    return "该字段来自官方 SDE 原始数据。"


def file_key_from_stem(stem: str) -> str:
    return normalize_key(stem.replace("__", "_").replace(".jsonl", ""))


def file_display(stem: str) -> dict[str, str]:
    key = file_key_from_stem(stem)
    info = FILE_INFO.get(key)
    if info:
        return {"label": with_original(info["label"], stem), "desc": info["desc"]}
    return {"label": stem, "desc": "官方 SDE 数据文件。"}


def pick_lang(value: Any, preferred: tuple[str, ...] = PREFERRED_LANGS) -> str | None:
    if isinstance(value, str):
        v = value.strip()
        return v or None
    if isinstance(value, dict):
        lowered = {str(k).lower(): v for k, v in value.items()}
        for lang in preferred:
            candidate = lowered.get(lang)
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        for candidate in value.values():
            found = pick_lang(candidate, preferred)
            if found:
                return found
    if isinstance(value, list):
        for candidate in value:
            found = pick_lang(candidate, preferred)
            if found:
                return found
    return None


def pick_zh(value: Any) -> str | None:
    return pick_lang(value, ("zh", "zh-cn"))


def pick_en(value: Any) -> str | None:
    return pick_lang(value, ("en", "en-us"))


def localized_name(zh: str | None, en: str | None, fallback: str) -> str:
    title = zh or en or fallback
    if zh and en and zh != en:
        return f"{zh}（{en}）"
    return title


def flatten_text(value: Any) -> list[str]:
    if isinstance(value, str):
        value = value.strip()
        return [value] if value else []
    if isinstance(value, dict):
        out: list[str] = []
        for v in value.values():
            out.extend(flatten_text(v))
        return out
    if isinstance(value, list):
        out: list[str] = []
        for v in value:
            out.extend(flatten_text(v))
        return out
    return []


def safe_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return pick_lang(value)


def clean_html(value: str | None) -> str | None:
    if not value:
        return None
    text_value = re.sub(r"<a\s+href=showinfo:(\d+)>", "", value, flags=re.I)
    text_value = re.sub(r"</a>", "", text_value, flags=re.I)
    text_value = re.sub(r"<[^>]+>", "", text_value)
    text_value = unescape(text_value)
    text_value = re.sub(r"\s+", " ", text_value).strip()
    return text_value or None


def as_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text_value = str(value).strip()
    if not text_value:
        return None
    if re.fullmatch(r"\d+\.0", text_value):
        return text_value[:-2]
    return text_value


def infer_key(record: dict[str, Any], stem: str) -> str:
    if "_key" in record:
        return str(record["_key"])
    for key, value in record.items():
        if key.lower().endswith("id") and not isinstance(value, (dict, list)):
            return str(value)
    return f"{stem}-{abs(hash(json.dumps(record, ensure_ascii=False, sort_keys=True))) % 10**12}"


def guess_entity_kind(record: dict[str, Any], file_key: str) -> str | None:
    if file_key in ENTITY_FILE_HINTS:
        return ENTITY_FILE_HINTS[file_key]
    for field, kind in ID_FIELD_ENTITY_HINTS.items():
        if field in record:
            return kind
    return None


def infer_titles(record: dict[str, Any], stem: str, key: str, translation_map: dict[str, dict[str, str]]) -> tuple[str, str | None]:
    file_key = file_key_from_stem(stem)
    entity_kind = guess_entity_kind(record, file_key)
    entity_id = None
    for field, kind in ID_FIELD_ENTITY_HINTS.items():
        if kind == entity_kind and field in record and not isinstance(record[field], (dict, list)):
            entity_id = str(record[field])
            break
    if entity_id is None and entity_kind and "_key" in record and not isinstance(record["_key"], (dict, list)):
        entity_id = str(record["_key"])

    zh: str | None = None
    for field in ("name", "displayName", "description"):
        if field in record:
            zh = pick_zh(record[field])
            if zh:
                break
    if not zh and entity_kind and entity_id:
        zh = translation_map.get(entity_kind, {}).get(entity_id)

    en: str | None = None
    for field in ("name", "displayName", "description", "effectName", "iconFile"):
        if field in record:
            if not zh:
                zh = pick_zh(record[field])
            en = pick_en(record[field]) if field != "effectName" else safe_text(record[field])
            if en or zh:
                break

    if not zh and "effectName" in record:
        zh = safe_text(record["effectName"])
    if not en and "effectName" in record:
        en = safe_text(record["effectName"])
    if not zh and "iconFile" in record:
        zh = safe_text(record["iconFile"])
    if not en and "iconFile" in record:
        en = safe_text(record["iconFile"])

    title = localized_name(zh, en, f"{stem} #{key}")
    alt = en if en and en != title else None
    return title, alt


def collect_search_tokens(record: dict[str, Any], stem: str, title: str, alt_title: str | None) -> str:
    tokens = [stem, title]
    if alt_title:
        tokens.append(alt_title)
    ui = record.get("_ui") if isinstance(record.get("_ui"), dict) else {}
    market_path = ui.get("marketPath")
    if isinstance(market_path, list):
        tokens.extend(str(part) for part in market_path)
    for field, value in record.items():
        if field.lower().endswith("id") and not isinstance(value, (dict, list)):
            tokens.append(str(value))
    return " ".join(" ".join(tokens).lower().split())


def short_value(value: Any) -> str:
    if isinstance(value, dict):
        zh = pick_zh(value)
        en = pick_en(value)
        if zh and en and zh != en:
            return f"{zh} / {en}"
        return zh or en or "对象"
    if isinstance(value, list):
        return f"{len(value)} 项"
    if isinstance(value, bool):
        return "是" if value else "否"
    text_value = safe_text(value)
    if text_value is None:
        return "空"
    if len(text_value) > 80:
        return text_value[:77] + "..."
    return text_value


def build_field_notes(record: dict[str, Any]) -> list[dict[str, str]]:
    notes: list[dict[str, str]] = []
    for field in record.keys():
        if field.startswith("_"):
            continue
        notes.append(
            {
                "key": field,
                "label": humanize_field(field),
                "meaning": field_meaning(field),
                "preview": short_value(record[field]),
            }
        )
    notes.sort(key=lambda item: (0 if item["key"] in FIELD_INFO else 1, item["label"]))
    return notes[:18]


def infer_summary(record: dict[str, Any], stem: str, key: str, title: str, alt_title: str | None) -> str:
    file_info = file_display(stem)
    bits = [file_info["label"], f"键值 {key}"]
    if alt_title:
        bits.append(alt_title)
    ui = record.get("_ui") if isinstance(record.get("_ui"), dict) else {}
    market_path = ui.get("marketPath")
    if isinstance(market_path, list) and market_path:
        bits.append("市场: " + " / ".join(str(part) for part in market_path[:5]))
    for field in ("groupID", "categoryID", "marketGroupID", "metaGroupID", "factionID", "raceID", "published"):
        if field in record:
            bits.append(f"{humanize_field(field)}: {short_value(record[field])}")
    return " · ".join(bits)


def compact(record: dict[str, Any], stem: str, translation_map: dict[str, dict[str, str]]) -> dict[str, Any]:
    out = dict(record)
    key = infer_key(out, stem)
    title, alt_title = infer_titles(out, stem, key, translation_map)
    file_info = file_display(stem)
    file_key = file_key_from_stem(stem)
    entity_kind = guess_entity_kind(out, file_key)
    entity_id = key
    market_path = None
    localized_description = translation_map.get("type_description", {}).get(entity_id) if entity_kind == "type" else None
    location_meta = None
    if entity_kind == "system":
        location_meta = translation_map.get("system_meta", {}).get(entity_id)
    elif entity_kind == "station":
        location_meta = translation_map.get("station_meta", {}).get(entity_id)
    out["_ui"] = {
        "key": key,
        "title": title,
        "altTitle": alt_title,
        "marketPath": market_path,
        "localizedDescription": localized_description,
        "location": location_meta,
        "summary": "",
        "fileLabel": file_info["label"],
        "fileDesc": file_info["desc"],
        "fieldNotes": build_field_notes(out),
    }
    out["_ui"]["summary"] = infer_summary(out, stem, key, title, alt_title)
    return out


def process_file(
    source: Path,
    target_root: Path,
    search_entries: list[dict[str, Any]],
    rel_name: str,
    translation_map: dict[str, dict[str, str]],
) -> dict[str, Any]:
    stem = rel_name.replace("/", "__").removesuffix(".jsonl")
    out_dir = target_root / stem
    out_dir.mkdir(parents=True, exist_ok=True)
    shard_i = 0
    shard: list[dict[str, Any]] = []
    shard_bytes = 0
    total = 0
    fields: Counter[str] = Counter()

    def flush() -> None:
        nonlocal shard_i, shard, shard_bytes
        if not shard:
            return
        (out_dir / f"{shard_i}.json").write_text(
            json.dumps(shard, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        shard_i += 1
        shard = []
        shard_bytes = 0

    with source.open("r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            record = json.loads(line)
            if not isinstance(record, dict):
                record = {"_value": record}
            total += 1
            fields.update(record.keys())
            item = compact(record, stem, translation_map)
            ui = item["_ui"]
            if stem in SEARCH_INDEX_FILES:
                search_entries.append(
                    {
                        "file": stem,
                        "key": ui["key"],
                        "title": ui["title"],
                        "altTitle": ui["altTitle"],
                        "summary": ui["summary"],
                        "fileLabel": ui["fileLabel"],
                        "q": collect_search_tokens(item, stem, ui["title"], ui["altTitle"]),
                    }
                )
            size = len(json.dumps(item, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
            if shard and (len(shard) >= MAX_RECORDS or shard_bytes + size > MAX_BYTES):
                flush()
            shard.append(item)
            shard_bytes += size
    flush()

    file_info = file_display(stem)
    manifest = {
        "file": stem,
        "source": rel_name,
        "records": total,
        "shards": shard_i,
        "topFields": [k for k, _ in fields.most_common(16)],
        "title": file_info["label"],
        "description": file_info["desc"],
    }
    (out_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def xlsx_headers(row: list[Any]) -> list[str]:
    headers: list[str] = []
    for value in row:
        text_value = safe_text(value) or ""
        headers.append(text_value.strip())
    return headers


def header_index(headers: list[str], names: tuple[str, ...]) -> int | None:
    normalized_names = {normalize_key(name) for name in names if normalize_key(name) and normalize_key(name) != "id"}
    lowered_names = {name.lower() for name in names}
    for idx, header in enumerate(headers):
        lowered = header.lower()
        if lowered in lowered_names:
            return idx
    for idx, header in enumerate(headers):
        normalized = normalize_key(header)
        if normalized and normalized in normalized_names:
            return idx
    for idx, header in enumerate(headers):
        normalized = normalize_key(header)
        lowered = header.lower()
        if normalized and any(name and name in normalized for name in normalized_names):
            return idx
        if any(name and name in lowered for name in lowered_names):
            return idx
    return None


def guess_header_indexes(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    normalized = [normalize_key(h) for h in headers]
    id_idx = None
    zh_idx = None
    en_idx = None
    id_priority = [
        "typeid", "itemid", "regionid", "constellationid", "solarsystemid", "systemid",
        "stationid", "npcstationid", "groupid", "categoryid", "marketgroupid", "factionid", "raceid", "id",
    ]
    zh_tokens = ("中文", "cn", "zh", "chinese", "namecn", "namezh", "cnname", "zhname", "名称")
    en_tokens = ("english", "nameen", "enname", "en", "英文")

    for token in id_priority:
        for idx, header in enumerate(normalized):
            if header == token or header.endswith(token):
                id_idx = idx
                break
        if id_idx is not None:
            break
    if id_idx is None:
        id_idx = header_index(
            headers,
            (
                "物品ID", "星域ID", "星座ID", "星系ID", "空间站ID", "建筑物ID",
                "分组ID", "大类ID", "市场分类ID", "势力ID", "种族ID", "ID",
            ),
        )
    for idx, original in enumerate(headers):
        if any(token in original.lower() for token in zh_tokens):
            zh_idx = idx
            break
    if zh_idx is None:
        zh_idx = header_index(headers, ("物品名称", "星域名字", "星座名字", "星系名字", "空间站名称", "建筑物名称", "名称", "名字"))
    for idx, original in enumerate(headers):
        if any(token in original.lower() for token in en_tokens):
            en_idx = idx
            break
    if zh_idx is None:
        for idx, header in enumerate(normalized):
            if header in {"name", "itemname"}:
                zh_idx = idx
                break
    return id_idx, zh_idx, en_idx


def guess_sheet_entity(sheet_name: str, headers: list[str]) -> str | None:
    normalized_sheet = normalize_key(sheet_name)
    if "物品" in sheet_name:
        return "type"
    if "空间站" in sheet_name or "建筑" in sheet_name:
        return "station"
    if "星系" in sheet_name:
        return "system"
    if "星座" in sheet_name:
        return "constellation"
    if "星域" in sheet_name:
        return "region"
    if "station" in normalized_sheet:
        return "station"
    if "solarsystem" in normalized_sheet or normalized_sheet.endswith("system"):
        return "system"
    if "constellation" in normalized_sheet:
        return "constellation"
    if "region" in normalized_sheet:
        return "region"
    if "item" in normalized_sheet or "type" in normalized_sheet:
        return "type"

    joined = " ".join(headers).lower()
    if "space station" in joined or "npc空间站" in joined or "stationid" in joined:
        return "station"
    if "solarsystemid" in joined or "systemid" in joined:
        return "system"
    if "constellationid" in joined:
        return "constellation"
    if "regionid" in joined:
        return "region"
    if "typeid" in joined or "itemid" in joined:
        return "type"
    return None


def build_translation_map(xlsx_path: Path | None) -> dict[str, dict[str, Any]]:
    mapping: dict[str, dict[str, Any]] = {
        kind: {}
        for kind in {
            "type",
            "region",
            "constellation",
            "system",
            "station",
            "group",
            "category",
            "marketgroup",
            "faction",
            "race",
            "type_description",
            "region_meta",
            "constellation_meta",
            "system_meta",
            "station_meta",
        }
    }
    if not xlsx_path or not xlsx_path.exists() or load_workbook is None:
        return mapping

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        headers = xlsx_headers(list(header_row))
        id_idx, zh_idx, en_idx = guess_header_indexes(headers)
        kind = guess_sheet_entity(sheet.title, headers)
        if id_idx is None or zh_idx is None or kind is None:
            continue

        for row in rows:
            row = list(row)
            if id_idx >= len(row):
                continue
            raw_id = row[id_idx]
            if raw_id is None:
                continue
            text_id = as_id(raw_id)
            if not text_id or text_id.lower() == "id":
                continue
            zh_value = safe_text(row[zh_idx]) if zh_idx < len(row) else None
            if not zh_value:
                continue
            mapping[kind][text_id] = zh_value
            if en_idx is not None and en_idx < len(row):
                en_value = safe_text(row[en_idx])
                if en_value and kind == "type":
                    mapping.setdefault("type_en", {})[text_id] = en_value

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        headers = xlsx_headers(list(header_row))
        title = sheet.title

        if "物品" in title:
            id_idx = header_index(headers, ("typeID", "物品ID"))
            name_idx = header_index(headers, ("物品名称", "名称"))
            desc_idx = header_index(headers, ("描述",))
            for row in rows:
                row = list(row)
                if id_idx is None or id_idx >= len(row):
                    continue
                text_id = as_id(row[id_idx])
                if not text_id:
                    continue
                if name_idx is not None and name_idx < len(row):
                    name = safe_text(row[name_idx])
                    if name:
                        mapping["type"][text_id] = name
                if desc_idx is not None and desc_idx < len(row):
                    desc = clean_html(safe_text(row[desc_idx]))
                    if desc:
                        mapping["type_description"][text_id] = desc

        elif "星域" in title:
            region_idx = header_index(headers, ("星域ID", "regionID"))
            region_name_idx = header_index(headers, ("星域名字", "星域名称", "name"))
            for row in rows:
                row = list(row)
                if region_idx is None or region_idx >= len(row):
                    continue
                region_id = as_id(row[region_idx])
                if not region_id:
                    continue
                region_name = safe_text(row[region_name_idx]) if region_name_idx is not None and region_name_idx < len(row) else None
                if region_name:
                    mapping["region"][region_id] = region_name
                    mapping["region_meta"][region_id] = {"name": region_name}

        elif "星座" in title:
            const_idx = header_index(headers, ("星座ID", "constellationID"))
            const_name_idx = header_index(headers, ("星座名字", "星座名称"))
            region_idx = header_index(headers, ("星域ID", "regionID"))
            region_name_idx = header_index(headers, ("星域名字", "星域名称"))
            for row in rows:
                row = list(row)
                if const_idx is None or const_idx >= len(row):
                    continue
                const_id = as_id(row[const_idx])
                if not const_id:
                    continue
                const_name = safe_text(row[const_name_idx]) if const_name_idx is not None and const_name_idx < len(row) else None
                region_id = as_id(row[region_idx]) if region_idx is not None and region_idx < len(row) else None
                region_name = safe_text(row[region_name_idx]) if region_name_idx is not None and region_name_idx < len(row) else None
                if const_name:
                    mapping["constellation"][const_id] = const_name
                mapping["constellation_meta"][const_id] = {
                    "name": const_name,
                    "regionID": region_id,
                    "regionName": region_name,
                }

        elif "星系" in title:
            system_idx = header_index(headers, ("星系ID", "solarSystemID", "systemID"))
            system_name_idx = header_index(headers, ("星系名字", "星系名称"))
            const_idx = header_index(headers, ("星座ID", "constellationID"))
            const_name_idx = header_index(headers, ("星座名字", "星座名称"))
            region_idx = header_index(headers, ("星域ID", "regionID"))
            region_name_idx = header_index(headers, ("星域名字", "星域名称"))
            security_idx = header_index(headers, ("安全等级", "securityStatus"))
            for row in rows:
                row = list(row)
                if system_idx is None or system_idx >= len(row):
                    continue
                system_id = as_id(row[system_idx])
                if not system_id:
                    continue
                system_name = safe_text(row[system_name_idx]) if system_name_idx is not None and system_name_idx < len(row) else None
                const_id = as_id(row[const_idx]) if const_idx is not None and const_idx < len(row) else None
                const_name = safe_text(row[const_name_idx]) if const_name_idx is not None and const_name_idx < len(row) else None
                region_id = as_id(row[region_idx]) if region_idx is not None and region_idx < len(row) else None
                region_name = safe_text(row[region_name_idx]) if region_name_idx is not None and region_name_idx < len(row) else None
                security = row[security_idx] if security_idx is not None and security_idx < len(row) else None
                if system_name:
                    mapping["system"][system_id] = system_name
                mapping["system_meta"][system_id] = {
                    "name": system_name,
                    "constellationID": const_id,
                    "constellationName": const_name,
                    "regionID": region_id,
                    "regionName": region_name,
                    "security": security,
                }

        elif "NPC空间站" in title:
            station_idx = header_index(headers, ("空间站ID", "stationID"))
            station_name_idx = header_index(headers, ("空间站名称", "空间站名字"))
            system_idx = header_index(headers, ("星系ID", "solarSystemID", "systemID"))
            system_name_idx = header_index(headers, ("星系名字", "星系名称"))
            const_idx = header_index(headers, ("星座ID", "constellationID"))
            const_name_idx = header_index(headers, ("星座名字", "星座名称"))
            region_idx = header_index(headers, ("星域ID", "regionID"))
            region_name_idx = header_index(headers, ("星域名字", "星域名称"))
            security_idx = header_index(headers, ("安全等级", "securityStatus"))
            for row in rows:
                row = list(row)
                if station_idx is None or station_idx >= len(row):
                    continue
                station_id = as_id(row[station_idx])
                if not station_id:
                    continue
                station_name = safe_text(row[station_name_idx]) if station_name_idx is not None and station_name_idx < len(row) else None
                system_id = as_id(row[system_idx]) if system_idx is not None and system_idx < len(row) else None
                system_name = safe_text(row[system_name_idx]) if system_name_idx is not None and system_name_idx < len(row) else None
                const_id = as_id(row[const_idx]) if const_idx is not None and const_idx < len(row) else None
                const_name = safe_text(row[const_name_idx]) if const_name_idx is not None and const_name_idx < len(row) else None
                region_id = as_id(row[region_idx]) if region_idx is not None and region_idx < len(row) else None
                region_name = safe_text(row[region_name_idx]) if region_name_idx is not None and region_name_idx < len(row) else None
                security = row[security_idx] if security_idx is not None and security_idx < len(row) else None
                if station_name:
                    mapping["station"][station_id] = station_name
                mapping["station_meta"][station_id] = {
                    "name": station_name,
                    "systemID": system_id,
                    "systemName": system_name,
                    "constellationID": const_id,
                    "constellationName": const_name,
                    "regionID": region_id,
                    "regionName": region_name,
                    "security": security,
                }

    return mapping


def iter_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            obj = json.loads(line)
            if isinstance(obj, dict):
                records.append(obj)
    return records


def entity_name(record: dict[str, Any], kind: str, mapping: dict[str, dict[str, Any]]) -> tuple[str, str | None]:
    key = str(record.get("_key") or record.get(f"{kind}ID") or "")
    zh = pick_zh(record.get("name")) or pick_zh(record.get("displayName"))
    if not zh:
        zh = mapping.get(kind, {}).get(key)
    en = pick_en(record.get("name")) or pick_en(record.get("displayName"))
    title = str(localized_name(zh, en, key))
    return title, en if en and en != title else None


def security_band(value: Any) -> str:
    try:
        sec = float(value)
    except (TypeError, ValueError):
        return "未知安全等级"
    if sec >= 0.45:
        return "高安"
    if sec > 0.0:
        return "低安"
    return "零安/虫洞"


def add_market_path(root: dict[str, Any], path: list[str], type_id: str) -> None:
    node = root
    for name in path:
        children = node.setdefault("children", {})
        node = children.setdefault(name, {"name": name, "children": {}, "typeIDs": []})
    node.setdefault("typeIDs", []).append(type_id)


def compact_tree(node: dict[str, Any]) -> dict[str, Any]:
    children = node.get("children", {})
    return {
        "name": node.get("name", "市场"),
        "typeIDs": sorted(set(node.get("typeIDs", [])), key=lambda value: int(value) if str(value).isdigit() else str(value)),
        "children": [compact_tree(child) for _, child in sorted(children.items(), key=lambda item: item[0])],
    }


def build_game_data(extracted: Path, mapping: dict[str, dict[str, Any]]) -> dict[str, Any]:
    types = iter_jsonl(extracted / "types.jsonl")
    groups = {str(item.get("_key")): item for item in iter_jsonl(extracted / "groups.jsonl")}
    categories = {str(item.get("_key")): item for item in iter_jsonl(extracted / "categories.jsonl")}
    market_groups = {str(item.get("_key")): item for item in iter_jsonl(extracted / "marketGroups.jsonl")}
    regions = {str(item.get("_key")): item for item in iter_jsonl(extracted / "mapRegions.jsonl")}
    constellations = {str(item.get("_key")): item for item in iter_jsonl(extracted / "mapConstellations.jsonl")}
    systems = {str(item.get("_key")): item for item in iter_jsonl(extracted / "mapSolarSystems.jsonl")}
    stations = iter_jsonl(extracted / "npcStations.jsonl")

    type_index: dict[str, dict[str, Any]] = {}
    market_nodes: dict[str, dict[str, Any]] = {"root": {"id": "root", "name": "市场", "children": [], "typeIDs": []}}
    for group_id, record in market_groups.items():
        name, _ = entity_name(record, "marketgroup", mapping)
        market_nodes[group_id] = {
            "id": group_id,
            "name": name,
            "parent": as_id(record.get("parentGroupID")),
            "children": [],
            "typeIDs": [],
        }
    for group_id, node in market_nodes.items():
        if group_id == "root":
            continue
        parent_id = node.get("parent")
        parent = market_nodes.get(parent_id or "root", market_nodes["root"])
        parent["children"].append(group_id)

    def market_path_for(group_id: str | None) -> list[str]:
        if not group_id:
            return []
        path: list[str] = []
        current_id = group_id
        seen: set[str] = set()
        while current_id and current_id not in seen:
            seen.add(current_id)
            node = market_nodes.get(current_id)
            if not node:
                break
            path.insert(0, str(node["name"]))
            current_id = node.get("parent")
        return path

    for record in types:
        type_id = as_id(record.get("_key"))
        if not type_id:
            continue
        title, alt = entity_name(record, "type", mapping)
        group_id = as_id(record.get("groupID"))
        group = groups.get(group_id or "")
        category = categories.get(str(group.get("categoryID")) if group else "")
        market_group_id = as_id(record.get("marketGroupID"))
        market_path = market_path_for(market_group_id)
        if market_group_id and market_group_id in market_nodes:
            market_nodes[market_group_id]["typeIDs"].append(type_id)
        group_name, _ = entity_name(group, "group", mapping) if group else (None, None)
        category_name, _ = entity_name(category, "category", mapping) if category else (None, None)
        desc = clean_html(pick_zh(record.get("description"))) or mapping.get("type_description", {}).get(type_id)
        type_index[type_id] = {
            "id": type_id,
            "name": title,
            "en": alt,
            "published": bool(record.get("published")),
            "groupID": group_id,
            "groupName": group_name,
            "categoryID": as_id(group.get("categoryID")) if group else None,
            "categoryName": category_name,
            "marketGroupID": market_group_id,
            "marketPath": market_path or [],
            "description": desc,
        }

    def compact_market_node(node_id: str) -> dict[str, Any]:
        node = market_nodes[node_id]
        children = [compact_market_node(child_id) for child_id in node["children"]]
        children.sort(key=lambda item: item["name"] or "")
        return {
            "id": node["id"],
            "name": node["name"],
            "typeIDs": sorted(set(node["typeIDs"]), key=lambda value: int(value) if str(value).isdigit() else str(value)),
            "children": children,
        }

    region_tree: dict[str, dict[str, Any]] = {}
    for region_id, record in regions.items():
        name, _ = entity_name(record, "region", mapping)
        region_tree[region_id] = {"id": region_id, "name": name, "constellations": {}}
    for const_id, record in constellations.items():
        region_id = as_id(record.get("regionID"))
        if not region_id:
            continue
        name, _ = entity_name(record, "constellation", mapping)
        region = region_tree.setdefault(region_id, {"id": region_id, "name": region_id, "constellations": {}})
        region["constellations"][const_id] = {"id": const_id, "name": name, "systems": {}}
    for system_id, record in systems.items():
        region_id = as_id(record.get("regionID"))
        const_id = as_id(record.get("constellationID"))
        if not region_id or not const_id:
            continue
        name, _ = entity_name(record, "system", mapping)
        region = region_tree.setdefault(region_id, {"id": region_id, "name": region_id, "constellations": {}})
        const = region["constellations"].setdefault(const_id, {"id": const_id, "name": const_id, "systems": {}})
        const["systems"][system_id] = {
            "id": system_id,
            "name": name,
            "security": record.get("securityStatus"),
            "securityBand": security_band(record.get("securityStatus")),
            "stations": [],
        }
    for record in stations:
        station_id = as_id(record.get("_key"))
        system_id = as_id(record.get("solarSystemID"))
        station_name = mapping.get("station", {}).get(station_id or "") or f"NPC 空间站 {station_id}"
        system_record = systems.get(system_id or "")
        if not station_id or not system_id or not system_record:
            continue
        region_id = as_id(system_record.get("regionID"))
        const_id = as_id(system_record.get("constellationID"))
        if not region_id or not const_id:
            continue
        region = region_tree.setdefault(region_id, {"id": region_id, "name": region_id, "constellations": {}})
        const = region["constellations"].setdefault(const_id, {"id": const_id, "name": const_id, "systems": {}})
        system = const["systems"].setdefault(system_id, {"id": system_id, "name": system_id, "stations": []})
        type_id = as_id(record.get("typeID"))
        station_type = type_index.get(type_id or "", {}).get("name")
        system.setdefault("stations", []).append({"id": station_id, "name": station_name, "type": station_type})

    universe = []
    for region in sorted(region_tree.values(), key=lambda item: item.get("name") or ""):
        constellations = []
        for const in sorted(region["constellations"].values(), key=lambda item: item.get("name") or ""):
            systems = []
            for system in sorted(const["systems"].values(), key=lambda item: item.get("name") or ""):
                system["stations"] = sorted(system.get("stations", []), key=lambda item: item.get("name") or "")
                systems.append(system)
            constellations.append({"id": const["id"], "name": const["name"], "systems": systems})
        universe.append({"id": region["id"], "name": region["name"], "constellations": constellations})

    return {
        "typeIndex": type_index,
        "marketTree": compact_market_node("root"),
        "universe": universe,
        "counts": {
            "types": len(type_index),
            "marketTypes": sum(1 for item in type_index.values() if item.get("marketPath")),
            "regions": len(universe),
            "stations": len(stations),
        },
    }


def build(extracted: Path, build_no: int, variant: str, translation_map: dict[str, dict[str, str]], translation_meta: dict[str, Any]) -> None:
    clean(DIST)
    data_root = DIST / "data" / "files"
    data_root.mkdir(parents=True, exist_ok=True)
    search_entries: list[dict[str, Any]] = []
    manifests: list[dict[str, Any]] = []
    game_data = build_game_data(extracted, translation_map)

    files = sorted(extracted.rglob("*.jsonl"), key=lambda p: str(p.relative_to(extracted)).lower())
    if not files:
        raise RuntimeError("No JSONL files found in extracted SDE archive")

    for file in files:
        manifests.append(
            process_file(
                file,
                data_root,
                search_entries,
                file.relative_to(extracted).as_posix(),
                translation_map,
            )
        )

    search_entries.sort(key=lambda x: (x["title"].lower(), x["file"], x["key"]))

    glossary = [
        {"key": key, "label": value["label"], "meaning": value["meaning"]}
        for key, value in sorted(FIELD_INFO.items(), key=lambda item: item[1]["label"])
    ]

    (DIST / "data" / "search-index.json").write_text(
        json.dumps({"minimumQueryLength": SEARCH_MIN, "entries": search_entries}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (DIST / "data" / "field-glossary.json").write_text(json.dumps(glossary, ensure_ascii=False, indent=2), encoding="utf-8")
    (DIST / "data" / "game-data.json").write_text(
        json.dumps(game_data, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (DIST / "data" / "meta.json").write_text(
        json.dumps(
            {
                "siteTitle": "EVE SDE 中文资料站",
                "buildNumber": build_no,
                "variant": variant,
                "generatedAt": datetime.now(timezone.utc).isoformat(),
                "fileCount": len(manifests),
                "gameDataCounts": game_data["counts"],
                "files": manifests,
                "translationMeta": translation_meta,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    for path in SRC.rglob("*"):
        if path.is_file():
            target = DIST / path.relative_to(SRC)
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, target)


def main() -> None:
    TMP.mkdir(parents=True, exist_ok=True)
    build_no = latest_build()
    zip_url = ZIP_TEMPLATE.format(build=build_no, variant=VARIANT)

    translation_meta: dict[str, Any] = {
        "displayLang": "zh",
        "sdeLocalizedText": "优先使用官方 SDE 自带中文字段，其次回退英文。",
        "externalTranslationWorkbook": None,
    }

    with tempfile.TemporaryDirectory(dir=TMP) as tmp_name:
        tmp = Path(tmp_name)
        zip_path = tmp / f"eve-online-static-data-{build_no}-{VARIANT}.zip"
        extracted = tmp / "extracted"
        extracted.mkdir(parents=True, exist_ok=True)

        print(f"Downloading official SDE build {build_no} from {zip_url}")
        download(zip_url, zip_path)

        print(f"Extracting {zip_path.name}")
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(extracted)

        translation_map: dict[str, dict[str, str]] = {}
        ceve_path = tmp / "evedata.xlsx"
        try:
            if CEVE_XLSX_URL and load_workbook is not None:
                print(f"Downloading CEVE translation workbook from {CEVE_XLSX_URL}")
                download(CEVE_XLSX_URL, ceve_path)
                translation_map = build_translation_map(ceve_path)
                translation_meta["externalTranslationWorkbook"] = {
                    "url": CEVE_XLSX_URL,
                    "status": "loaded",
                    "sheetsApplied": sorted(k for k, v in translation_map.items() if v and not k.endswith("_meta")),
                }
            else:
                translation_meta["externalTranslationWorkbook"] = {
                    "url": CEVE_XLSX_URL,
                    "status": "skipped",
                    "reason": "openpyxl unavailable",
                }
        except Exception as exc:
            translation_meta["externalTranslationWorkbook"] = {
                "url": CEVE_XLSX_URL,
                "status": "failed",
                "reason": str(exc),
            }
            print(f"Warning: failed to load CEVE workbook: {exc}")

        print("Building localized static site")
        build(extracted, build_no, VARIANT, translation_map, translation_meta)
        print(f"Done. Output written to {DIST}")


if __name__ == "__main__":
    main()
